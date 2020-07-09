# 用来读取测试数据
from openpyxl import load_workbook
# 变成函数
def read_data(file_name,sheet_name):#读取数据的函数
    wb=load_workbook(file_name)
    sheet=wb[sheet_name]
    all_case=[]#存储所有行的测试用例数据
    for i in range(2,sheet.max_row+1):
        case=[]#某一行的测试用例数据
        for j in range(1,sheet.max_column-1):
            case.append(sheet.cell(row=i,column=j).value)
            # print(sheet.cell(row=2,column=j).value)
        # print(case)
        all_case.append(case)
    return all_case
    # print(all_case)
if __name__ == '__main__':#只有在当前文件下才会执行
    all_case=read_data('test_data.xlsx','recharge')
    print('所有的测试数据为：',all_case)